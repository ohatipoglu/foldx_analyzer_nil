"""
Excel export module for FoldX analysis results.

Handles exporting analysis results to multi-sheet Excel files with:
- Summary sheet with statistics and decisions
- Detailed sheet with all calculated ΔΔG values
- Batch summary for multi-file processing
"""

import os
from typing import List, Optional, Dict, Any
import pandas as pd
from datetime import datetime

from utils.config import (
    SUMMARY_SHEET_NAME,
    DETAILED_SHEET_NAME,
    OUTPUT_FILE_SUFFIX,
    OUTPUT_EXCELS_DIR,
)
from utils.logger import get_logger

logger = get_logger()


class ExcelExportError(Exception):
    """Exception raised for Excel export errors."""
    pass


class FoldXExporter:
    """
    Exporter for FoldX analysis results to Excel format.
    
    Features:
    - Multi-sheet Excel files (Summary + Detailed)
    - Formatted output with column styling
    - Batch summary for multiple files
    - Error handling for open files
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize exporter.
        
        Args:
            output_dir: Directory for output files (default: same as input)
        """
        self.output_dir = output_dir
        self.logger = get_logger()
    
    def export_analysis(
        self,
        summary_df: pd.DataFrame,
        detailed_df: pd.DataFrame,
        input_filepath: str,
        custom_filename: Optional[str] = None
    ) -> str:
        """
        Export analysis results to Excel file.
        
        Args:
            summary_df: Summary statistics DataFrame
            detailed_df: Detailed ΔΔG calculations DataFrame
            input_filepath: Path to original .fxout file
            custom_filename: Optional custom output filename
            
        Returns:
            Path to created Excel file
            
        Raises:
            ExcelExportError: If export fails
        """
        # Determine output path
        output_path = self._get_output_path(
            input_filepath, custom_filename
        )
        
        self.logger.info(f"Exporting results to: {os.path.basename(output_path)}")
        
        try:
            # Check if file is already open
            if self._is_file_open(output_path):
                raise ExcelExportError(
                    f"Excel file is open in another program: {output_path}"
                )
            
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Summary Analysis
                self._write_summary_sheet(summary_df, writer)
                
                # Sheet 2: Detailed Calculations
                self._write_detailed_sheet(detailed_df, writer)
            
            self.logger.success(f"Excel file created successfully")
            return output_path
            
        except PermissionError:
            raise ExcelExportError(
                f"Permission denied - file may be open: {output_path}"
            )
        except Exception as e:
            raise ExcelExportError(f"Export failed: {str(e)}")
    
    def _get_output_path(
        self,
        input_filepath: str,
        custom_filename: Optional[str] = None
    ) -> str:
        """
        Determine the output Excel file path.

        Args:
            input_filepath: Path to input .fxout file
            custom_filename: Optional custom filename

        Returns:
            Full output path
        """
        input_basename = os.path.splitext(os.path.basename(input_filepath))[0]

        if custom_filename:
            filename = custom_filename
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
        else:
            filename = f"{input_basename}{OUTPUT_FILE_SUFFIX}"

        # Use configured output directory or input file's directory
        output_dir = self.output_dir if self.output_dir else OUTPUT_EXCELS_DIR

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        return os.path.join(output_dir, filename)
    
    def _write_summary_sheet(
        self,
        df: pd.DataFrame,
        writer: pd.ExcelWriter
    ) -> None:
        """
        Write summary statistics sheet with formatting.

        Args:
            df: Summary DataFrame
            writer: Excel writer object
        """
        # Base columns for display
        display_columns = [
            'AminoAcid', 'Valid_Runs', 'MEAN_ddG', 'STDEV_ddG',
            'SEM_ddG', 'MIN_ddG', 'MAX_ddG', 'Decision_Status',
            'Decision_Reason'
        ]
        
        # Add synthesizability columns if they exist (Biomatik filter)
        optional_columns = ['Synthesis_Status', 'Synthesis_Warnings']
        for col in optional_columns:
            if col in df.columns:
                display_columns.append(col)

        # Only include columns that exist
        available_columns = [c for c in display_columns if c in df.columns]
        export_df = df[available_columns].copy()

        # Round numeric columns for readability
        numeric_cols = ['MEAN_ddG', 'STDEV_ddG', 'SEM_ddG', 'MIN_ddG', 'MAX_ddG']
        for col in numeric_cols:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(4)

        # Write to Excel
        export_df.to_excel(
            writer,
            sheet_name=SUMMARY_SHEET_NAME,
            index=False,
            na_rep='N/A'
        )

        # Apply formatting
        self._format_summary_sheet(writer, SUMMARY_SHEET_NAME, len(export_df))

        self.logger.debug(
            f"Written {len(export_df)} rows to '{SUMMARY_SHEET_NAME}' sheet"
        )
    
    def _write_detailed_sheet(
        self,
        df: pd.DataFrame,
        writer: pd.ExcelWriter
    ) -> None:
        """
        Write detailed calculations sheet.
        
        Args:
            df: Detailed DataFrame
            writer: Excel writer object
        """
        # Round energy columns
        energy_cols = ['Mutant_Energy', 'WT_Energy', 'ddG']
        export_df = df.copy()
        
        for col in energy_cols:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(4)
        
        # Write to Excel
        export_df.to_excel(
            writer,
            sheet_name=DETAILED_SHEET_NAME,
            index=False,
            na_rep='N/A'
        )
        
        self.logger.debug(
            f"Written {len(export_df)} rows to '{DETAILED_SHEET_NAME}' sheet"
        )
    
    def _format_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        row_count: int
    ) -> None:
        """
        Apply formatting to summary sheet.

        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet
            row_count: Number of data rows
        """
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Get actual columns from the sheet
            columns_in_sheet = [cell.value for cell in worksheet[1]]

            # Base column widths
            column_widths = {
                'A': 12,  # AminoAcid
                'B': 12,  # Valid_Runs
                'C': 12,  # MEAN_ddG
                'D': 12,  # STDEV_ddG
                'E': 12,  # SEM_ddG
                'F': 12,  # MIN_ddG
                'G': 12,  # MAX_ddG
                'H': 28,  # Decision_Status
                'I': 40,  # Decision_Reason
            }
            
            # Add widths for optional synthesizability columns
            col_index = 9  # Start from J (after I)
            if 'Synthesis_Status' in columns_in_sheet:
                column_widths[chr(ord('A') + col_index)] = 18  # Synthesis_Status
                col_index += 1
            if 'Synthesis_Warnings' in columns_in_sheet:
                column_widths[chr(ord('A') + col_index)] = 50  # Synthesis_Warnings

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Header formatting
            from openpyxl.styles import Font, Alignment, PatternFill

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data alignment
            for row in worksheet.iter_rows(
                min_row=2, max_row=row_count + 1
            ):
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="center", 
                        vertical="center"
                    )
            
        except Exception as e:
            self.logger.warning(f"Could not apply formatting: {str(e)}")
    
    def _is_file_open(self, filepath: str) -> bool:
        """
        Check if Excel file is open in another program.
        
        Args:
            filepath: Path to check
            
        Returns:
            True if file appears to be open
        """
        try:
            # Try to open file in append mode
            with open(filepath, 'a') as f:
                pass
            return False
        except (PermissionError, IOError):
            return True
    
    def export_batch_summary(
        self,
        results: List[Dict[str, Any]],
        output_dir: str,
        filename: Optional[str] = None
    ) -> str:
        """
        Export batch processing summary to Excel.
        
        Args:
            results: List of result dictionaries from batch processing
            output_dir: Output directory
            filename: Optional custom filename
            
        Returns:
            Path to created Excel file
        """
        if not results:
            raise ExcelExportError("No results to export")
        
        # Create summary DataFrame
        summary_data = []
        for result in results:
            summary_data.append({
                'File': result.get('filename', 'Unknown'),
                'Status': result.get('status', 'Unknown'),
                'Ideal_Candidates': result.get('ideal_count', 0),
                'Rejected': result.get('rejected_count', 0),
                'Total_Analyzed': result.get('total_count', 0),
                'Best_ddG': result.get('best_ddG', None),
                'Worst_ddG': result.get('worst_ddG', None),
                'Output_File': result.get('output_path', 'N/A'),
                'Error': result.get('error', '')
            })
        
        batch_df = pd.DataFrame(summary_data)
        
        # Determine output path
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Batch_Summary_{timestamp}.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = os.path.join(output_dir, filename)
        
        self.logger.info(f"Exporting batch summary to: {filename}")
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                batch_df.to_excel(
                    writer,
                    sheet_name='Batch_Summary',
                    index=False
                )
            
            return output_path
            
        except Exception as e:
            raise ExcelExportError(f"Batch export failed: {str(e)}")


def export_to_excel(
    summary_df: pd.DataFrame,
    detailed_df: pd.DataFrame,
    input_filepath: str,
    output_dir: Optional[str] = None
) -> str:
    """
    Convenience function to export analysis results.
    
    Args:
        summary_df: Summary statistics
        detailed_df: Detailed calculations
        input_filepath: Original .fxout path
        output_dir: Optional output directory
        
    Returns:
        Path to created Excel file
    """
    exporter = FoldXExporter(output_dir=output_dir)
    return exporter.export_analysis(summary_df, detailed_df, input_filepath)
